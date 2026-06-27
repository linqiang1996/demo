from __future__ import annotations

import io
import re
from dataclasses import dataclass
from html import unescape
from pathlib import Path
from typing import Iterable, List

import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from .product_names import looks_like_product_name, normalize_product_name


DATE_COLUMNS = ["业务日期", "交易日期", "日期", "净值日期", "基金净值日期", "估值日期", "估值基准日", "date", "Date"]
NAV_COLUMNS = [
    "虚拟单位净值",
    "虚拟后净值",
    "虚拟净值",
    "试算单位净值（扣除业绩报酬后）",
    "试算单位净值(扣除业绩报酬后)",
    "单位净值",
    "产品总层面单位净值",
    "累计单位净值",
    "累计净值",
    "净值",
    "nav",
    "NAV",
]
PRODUCT_COLUMNS = ["产品名称", "基金名称", "基金全称", "产品", "基金", "产品全称", "name", "Name"]
HEADER_HINT_COLUMNS = PRODUCT_COLUMNS + DATE_COLUMNS + NAV_COLUMNS + ["基金代码", "客户代码", "客户名称", "发生份额", "实际提成金额"]


@dataclass
class ParsedProductNav:
    product_name: str
    records: list[tuple[str, float]]


PRODUCT_NAME_HINT_PATTERN = re.compile(
    r"[\u4e00-\u9fa5A-Za-z0-9]+(?:[\u4e00-\u9fa5A-Za-z0-9（）()·\-_]*)(?:私募证券投资基金|证券投资基金|私募基金|基金)[A-Za-z0-9]*"
)


def extract_product_name_candidates(*texts: object) -> list[str]:
    candidates: list[str] = []
    seen: set[str] = set()
    for value in texts:
        if value is None:
            continue
        text = normalize_product_name(str(value))
        if not text:
            continue
        split_texts = re.split(r"[_|｜]+", text)
        for segment in split_texts:
            cleaned_segment = normalize_product_name(segment)
            if not cleaned_segment:
                continue
            for match in PRODUCT_NAME_HINT_PATTERN.finditer(cleaned_segment):
                candidate = normalize_product_name(match.group(0))
                if candidate and candidate not in seen:
                    seen.add(candidate)
                    candidates.append(candidate)
            if looks_like_product_name(cleaned_segment) and cleaned_segment not in seen:
                seen.add(cleaned_segment)
                candidates.append(cleaned_segment)
        if looks_like_product_name(text) and text not in seen:
            seen.add(text)
            candidates.append(text)
    return candidates


def _normalize_header_cell(value: object) -> str:
    if value is None:
        return ""
    return normalize_product_name(str(value))


def _make_unique_columns(columns: list[str]) -> list[str]:
    counts: dict[str, int] = {}
    unique: list[str] = []
    for idx, column in enumerate(columns):
        base = column or f"__col_{idx}"
        count = counts.get(base, 0)
        unique_name = base if count == 0 else f"{base}__{count}"
        counts[base] = count + 1
        unique.append(unique_name)
    return unique


def _detect_header_row(df: pd.DataFrame) -> int | None:
    best_row: int | None = None
    best_score = 0
    header_candidates = {_normalize_header_cell(item) for item in HEADER_HINT_COLUMNS}
    for row_idx in range(min(len(df), 8)):
        values = [_normalize_header_cell(value) for value in df.iloc[row_idx].tolist()]
        score = sum(1 for value in values if value and value in header_candidates)
        if score > best_score:
            best_score = score
            best_row = row_idx
    if best_score >= 2:
        return best_row
    return None


def _read_sheet(source: object, sheet_name: str) -> pd.DataFrame:
    raw_df = pd.read_excel(source, sheet_name=sheet_name, header=None)
    if raw_df.empty:
        return raw_df
    header_row = _detect_header_row(raw_df)
    if header_row is None:
        return pd.read_excel(source, sheet_name=sheet_name)
    header_values = _make_unique_columns(
        [_normalize_header_cell(value) or f"__col_{idx}" for idx, value in enumerate(raw_df.iloc[header_row].tolist())]
    )
    body = raw_df.iloc[header_row + 1 :].copy()
    body.columns = header_values
    body = body.dropna(how="all")
    non_empty_mask = ~body.isna().all(axis=0)
    body = body.loc[:, non_empty_mask]
    return body.reset_index(drop=True)


def _normalize_table_text(value: object) -> str:
    if value is None:
        return ""
    return normalize_product_name(unescape(str(value)))

def _find_column(candidates: Iterable[str], columns: Iterable[str]) -> str | None:
    cleaned = {str(column).strip(): column for column in columns}
    for candidate in candidates:
        for column in columns:
            if str(column).strip().lower() == candidate.lower():
                return str(column)
        if candidate in cleaned:
            return cleaned[candidate]
    return None


def _coerce_date_series(series: pd.Series) -> pd.Series:
    text = series.astype(str).str.strip()
    parsed = pd.to_datetime(text, errors="coerce")
    missing = parsed.isna()
    if missing.any():
        eight_digits = text.str.fullmatch(r"\d{8}")
        reparsed = pd.to_datetime(text.where(eight_digits), format="%Y%m%d", errors="coerce")
        parsed = parsed.fillna(reparsed)
    return parsed


def _infer_product_name_from_values(values: Iterable[object], default_name: str) -> str:
    for candidate in extract_product_name_candidates(*list(values)):
        if looks_like_product_name(candidate):
            return candidate
    return default_name


def _extract_workbook_name_hints(content: bytes) -> list[str]:
    hints: list[str] = []
    try:
        workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception:
        return hints

    for sheet in workbook.worksheets[:3]:
        hints.extend(extract_product_name_candidates(sheet.title))
        for row in sheet.iter_rows(min_row=1, max_row=8, values_only=True):
            for value in row:
                if value in (None, ""):
                    continue
                hints.extend(extract_product_name_candidates(value))
    return [hint for hint in hints if hint]


def _resolve_default_name(default_name: str, extra_hints: Iterable[object] | None = None) -> str:
    hint_values: list[object] = [default_name]
    if extra_hints:
        hint_values.extend(list(extra_hints))
    for candidate in extract_product_name_candidates(*hint_values):
        if looks_like_product_name(candidate):
            return candidate
    return normalize_product_name(default_name)


def _coerce_nav_dataframe(df: pd.DataFrame, default_name: str) -> list[ParsedProductNav]:
    if df.empty:
        return []

    date_col = _find_column(DATE_COLUMNS, df.columns)
    nav_col = _find_column(NAV_COLUMNS, df.columns)
    product_col = _find_column(PRODUCT_COLUMNS, df.columns)

    if date_col is None or nav_col is None:
        return []

    working = df.copy()
    working = working[[column for column in [product_col, date_col, nav_col] if column is not None]]
    if product_col is None:
        inferred_name = _infer_product_name_from_values(df.iloc[:8].to_numpy().flatten().tolist(), default_name)
        working["__product_name__"] = inferred_name
        product_col = "__product_name__"

    working[date_col] = _coerce_date_series(working[date_col])
    working[nav_col] = pd.to_numeric(working[nav_col], errors="coerce")
    working[product_col] = working[product_col].astype(str).map(lambda value: normalize_product_name(value) or default_name)
    working = working.dropna(subset=[date_col, nav_col])

    parsed: list[ParsedProductNav] = []
    for product_name, group in working.groupby(product_col):
        if not looks_like_product_name(product_name):
            continue
        group = group.sort_values(date_col)
        records = [(timestamp.strftime("%Y-%m-%d"), float(nav)) for timestamp, nav in zip(group[date_col], group[nav_col])]
        if records:
            parsed.append(ParsedProductNav(product_name=product_name, records=records))
    return parsed


def parse_html_tables(body_text: str, extra_name_hints: Iterable[object] | None = None) -> list[ParsedProductNav]:
    if not body_text or "<table" not in body_text.lower():
        return []
    soup = BeautifulSoup(body_text, "html.parser")
    parsed: list[ParsedProductNav] = []
    default_name = _resolve_default_name("", extra_name_hints)
    for table in soup.find_all("table"):
        rows: list[list[str]] = []
        for tr in table.find_all("tr"):
            cells = tr.find_all(["th", "td"])
            row = [_normalize_table_text(cell.get_text(" ", strip=True)) for cell in cells]
            if any(row):
                rows.append(row)
        if not rows:
            continue
        width = max(len(row) for row in rows)
        normalized_rows = [row + [""] * (width - len(row)) for row in rows]
        df = pd.DataFrame(normalized_rows)
        parsed.extend(_coerce_nav_dataframe(df, default_name))
    return parsed


def parse_excel_bytes(content: bytes, filename: str, extra_name_hints: Iterable[object] | None = None) -> list[ParsedProductNav]:
    buffer = io.BytesIO(content)
    excel = pd.ExcelFile(buffer)
    default_name = _resolve_default_name(Path(filename).stem, extra_name_hints)
    name_hints = _extract_workbook_name_hints(content)
    if name_hints:
        default_name = _infer_product_name_from_values([default_name, *name_hints], default_name)
    parsed: list[ParsedProductNav] = []
    for sheet_name in excel.sheet_names:
        buffer.seek(0)
        df = _read_sheet(buffer, sheet_name)
        parsed.extend(_coerce_nav_dataframe(df, default_name))
    return parsed


def parse_excel_file(path: Path) -> list[ParsedProductNav]:
    default_name = normalize_product_name(path.stem)
    parsed: list[ParsedProductNav] = []
    excel = pd.ExcelFile(path)
    for sheet_name in excel.sheet_names:
        df = _read_sheet(path, sheet_name)
        parsed.extend(_coerce_nav_dataframe(df, default_name))
    return parsed


BODY_PATTERNS = [
    re.compile(
        r"(?P<name>[\u4e00-\u9fa5A-Za-z0-9（）()·\-_]+)[：:，,\s]+(?P<date>\d{4}[-/年]\d{1,2}[-/月]\d{1,2})[日]?[：:\s,，]+(?P<nav>\d+(?:\.\d+)?)"
    ),
    re.compile(
        r"(?P<date>\d{4}[-/年]\d{1,2}[-/月]\d{1,2})[日]?[，,\s]+(?P<name>[\u4e00-\u9fa5A-Za-z0-9（）()·\-_]+)[：:\s,，]+(?P<nav>\d+(?:\.\d+)?)"
    ),
]


def parse_body_text(body_text: str) -> list[ParsedProductNav]:
    grouped: dict[str, list[tuple[str, float]]] = {}
    for pattern in BODY_PATTERNS:
        for match in pattern.finditer(body_text):
            product_name = normalize_product_name(match.group("name"))
            date_text = (
                match.group("date")
                .replace("年", "-")
                .replace("月", "-")
                .replace("日", "")
                .replace("/", "-")
            )
            nav_value = float(match.group("nav"))
            try:
                nav_date = pd.to_datetime(date_text, errors="raise").strftime("%Y-%m-%d")
            except Exception:
                continue
            grouped.setdefault(product_name, []).append((nav_date, nav_value))

    parsed: list[ParsedProductNav] = []
    for product_name, records in grouped.items():
        if not looks_like_product_name(product_name):
            continue
        unique = sorted({record for record in records}, key=lambda item: item[0])
        parsed.append(ParsedProductNav(product_name=product_name, records=unique))
    if parsed:
        return parsed
    return parse_html_tables(body_text)
